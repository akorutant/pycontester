import os
import openpyxl
import datetime as dt

from flask import Flask, render_template, make_response, request, abort, url_for, flash, jsonify, send_from_directory
from flask_restful import abort, Api
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.utils import redirect

from data import db_session
from data.contests import Contest
from data.tasks import Task
from data.teachers import Teacher
from data.users import User
from data.contest_results import ContestResults
from forms.add_contest_form import AddContestForm
from forms.add_tasks_for_contest_form import AddTasksForContestForm
from forms.change_avatar_form import ChangeAvatarForm
from forms.change_password_form import ChangePasswordForm
from forms.feedback_form import FeedbackForm
from forms.login_form import LoginForm
from forms.register_form import RegisterForm
from forms.add_task_to_contest import AddTasksToContestForm
from tasks_api import TaskResource



app = Flask(__name__)
app.config["SECRET_KEY"] = "fjkFOEKFMOKMFIO3FMKLMkelfmOIJR3FMFKNFOU2IN3PIFNOI232F"
app.config["UPLOAD_FOLDER"] = "./static/files"

login_manager = LoginManager()
login_manager.init_app(app)

api = Api(app)
api.add_resource(TaskResource, '/api/task/<int:contest_id>')

if not os.path.isdir('database'):
    os.mkdir('database')

if not os.path.isfile('database/db.sqlite'):
    with open('database/db.sqlite', 'w') as f:
        pass

if not os.path.isdir("static/files"):
    os.mkdir("static/files")

db_session.global_init("database/db.sqlite")


@login_manager.user_loader
def load_user(user_id):
    db_sess = db_session.create_session()
    return db_sess.query(User).get(user_id)


@app.route('/')
@app.route('/main')
def index():
    return render_template('index.html',
                           title="Главная")


@app.route('/register', methods=['GET', 'POST'])
def register():
    form = RegisterForm()
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if form.validate_on_submit():
        if form.password.data != form.password_again.data:
            return render_template('register.html',
                                   title='Регистрация',
                                   form=form,
                                   message="Пароли не совпадают")
        db_sess = db_session.create_session()
        if db_sess.query(User).filter(User.email == form.email.data).first():
            return render_template('register.html',
                                   title='Регистрация',
                                   form=form,
                                   message="Такой пользователь уже есть")
        if form.select.data:
            user = User(
                login=form.login.data,
                firstname=form.firstname.data,
                surname=form.surname.data,
                patronymic=form.patronymic.data,
                job_title=form.job_title.data,
                email=form.email.data
            )

            user.set_password(form.password.data)
            db_sess.add(user)
            db_sess.commit()

            if form.job_title.data == "teacher":
                user = db_sess.query(User).filter_by(
                    email=form.email.data).first()
                teacher = Teacher(
                    user_id=user.id
                )
                db_sess.add(teacher)
                db_sess.commit()
            return redirect(url_for('login'))
    return render_template('register.html',
                           title='Регистрация',
                           form=form)


@app.route("/login", methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if form.validate_on_submit():
        db_sess = db_session.create_session()
        user = db_sess.query(User).filter(
            User.email == form.email.data).first()
        if user and user.check_password(form.password.data):
            login_user(user, remember=form.remember_me.data)
            return redirect(url_for("index"))
        return render_template('login.html',
                               message="Неправильный логин или пароль",
                               title='Авторизация',
                               form=form)
    return render_template('login.html',
                           title='Авторизация',
                           form=form)


@app.route("/account/<int:user_id>", methods=['GET', 'POST'])
@login_required
def account(user_id):
    if current_user.id == user_id:
        db_sess = db_session.create_session()
        user = db_sess.query(User).filter(User.id == current_user.id).first()
        form_change_password = ChangePasswordForm()
        form_change_avatar = ChangeAvatarForm()
        if form_change_avatar.validate_on_submit():
            message_for_avatar_form = "Вы не прикрепили файл"
            user_photo = request.files['avatar']
            if user_photo:
                user.avatar = user_photo.read()
                db_sess.commit()
                message_for_avatar_form = "Аватар обновлен"
            return render_template('personal_area.html',
                                   title="Аккаунт",
                                   message_for_avatar_form=message_for_avatar_form,
                                   form_change_password=form_change_password,
                                   form_change_avatar=form_change_avatar)

        elif form_change_password.validate_on_submit():
            message_for_password_form = "Неправильный пароль"
            if user and user.check_password(form_change_password.old_password.data):
                if form_change_password.new_password.data == form_change_password.repeated_new_password.data:
                    user.set_password(
                        form_change_password.repeated_new_password.data)
                    db_sess.commit()
                    message_for_password_form = "Пароль изменён"
                else:
                    message_for_password_form = "Пароли не совпадают"
            return render_template('personal_area.html',
                                   title="Аккаунт",
                                   message_for_password_form=message_for_password_form,
                                   form_change_password=form_change_password,
                                   form_change_avatar=form_change_avatar)
        return render_template('personal_area.html',
                               title="Аккаунт",
                               form_change_password=form_change_password,
                               form_change_avatar=form_change_avatar)

    return redirect(url_for("register"))


@app.route("/code")
@login_required
def code():
    return render_template('code.html',
                           title="Редактор кода")


#

@app.route("/contests")
@login_required
def contests():
    db_sess = db_session.create_session()
    contests_list = db_sess.query(Contest).all()
    contest_results = db_sess.query(ContestResults).filter(ContestResults.student_id == current_user.id).all()
    return render_template('contests.html',
                           title="Список соревнований",
                           time_now=dt.datetime.now(),
                           contests=contests_list,
                           contest_results=[i.contest_id for i in contest_results])


@app.route("/help")
def help():
    form = FeedbackForm()
    if form.validate_on_submit():
        return render_template('feedback.html',
                               title="Обратная связь",
                               form=form)
    return render_template('feedback.html',
                           title="Обратная связь",
                           form=form)


@app.route("/contests/<int:contest_id>", methods=['GET', 'POST'])
@login_required
def contests_list(contest_id):
    db_sess = db_session.create_session()
    contest = db_sess.query(Contest).filter(Contest.id == contest_id).first()

    contest_results = db_sess.query(ContestResults) \
        .filter(ContestResults.contest_id == contest_id) \
        .filter(current_user.id == ContestResults.student_id).first()

    if contest.end_deadline <= dt.datetime.now() or contest_results:
        return redirect(url_for("contests"))

    task_data = contest.tasks
    return render_template('contest_code.html',
                           title=contest.title,
                           contest=contest,
                           tasks=task_data
                           )


@app.route("/contests/teacher_list")
@login_required
def contests_teacher():
    if current_user.job_title == "teacher":
        db_sess = db_session.create_session()
        contests_data = db_sess.query(Contest).filter(
            Contest.author_id == current_user.id).all()
        return render_template("teacher_contests.html",
                               title="Список соревнований",
                               contests=contests_data,
                               now=dt.datetime.now())
    return redirect(url_for("index"))


@app.route("/contests/teacher_list/<int:id>", methods=['GET', 'POST'])
@login_required
def contests_edit(id):
    form = AddContestForm()
    if current_user.job_title == "teacher":
        if request.method == "GET":
            db_sess = db_session.create_session()
            contests_data = db_sess.query(Contest).filter(Contest.id == id,
                                                          Contest.author_id == current_user.id).first()
            if contests_data:
                form.contest_title.data = contests_data.title
                form.contest_description.data = contests_data.description
                form.join_deadline.data = contests_data.join_deadline
                form.end_deadline.data = contests_data.end_deadline
                form.submit.data = "Обновить"
            else:
                abort(404)
        if form.validate_on_submit():
            db_sess = db_session.create_session()
            contests_data = db_sess.query(Contest).filter(Contest.id == id,
                                                          Contest.author_id == current_user.id).first()
            if contests_data:
                contests_data.title = form.contest_title.data
                contests_data.description = form.contest_description.data
                contests_data.join_deadline = form.join_deadline.data
                contests_data.end_deadline = form.end_deadline.data
                db_sess.commit()
                return redirect(url_for("contests_teacher"))
            else:
                abort(404)
        return render_template("contests_add.html",
                               title="Редакторование конкурcов",
                               form=form,
                               current_datetime=dt.datetime.now())
    return redirect(url_for("index"))


@app.route("/contests/add", methods=["GET", "POST"])
@login_required
def contests_add():
    if current_user.job_title == "teacher":
        form = AddContestForm()
        if form.validate_on_submit():
            db_sess = db_session.create_session()
            contest = Contest(
                title=form.contest_title.data,
                description=form.contest_description.data,
                author_id=current_user.id,
                join_deadline=form.join_deadline.data,
                end_deadline=form.end_deadline.data
            )
            data = db_sess.query(Contest).filter(Contest.title == form.contest_title.data).first()
            if not data:
                db_sess.add(contest)
                db_sess.commit()
            return redirect(url_for("contests_teacher"))
        return render_template("contests_add.html",
                               title="Добавление конкурcов",
                               form=form,
                               current_datetime=dt.datetime.now())
    return redirect(url_for("index"))


@app.route("/contest_delete/<int:id>", methods=["GET", "POST"])
@login_required
def contest_delete(id):
    if current_user.job_title == "teacher":
        db_sess = db_session.create_session()
        contests_data = db_sess.query(Contest).filter(Contest.id == id,
                                                      Contest.author_id == current_user.id).first()
        if contests_data:
            db_sess.delete(contests_data)
            db_sess.commit()
        else:
            abort(404)
        return redirect(url_for("contests_teacher"))
    return redirect(url_for("index"))


@app.route("/contest/results/<int:contest_id>")
@login_required
def contests_results(contest_id):
    db_sess = db_session.create_session()
    contest = db_sess.query(Contest).filter(Contest.id == contest_id).first()
    contest_results = db_sess.query(ContestResults).filter(ContestResults.contest_id == contest_id)\
        .order_by(ContestResults.complited.desc()).all()
    return render_template('contest_results.html', contest=contest, results=contest_results)


@app.route("/tasks/<int:contest_id>", methods=['GET', 'POST'])
@login_required
def tasks(contest_id):
    if current_user.job_title == "teacher":
        form = AddTasksToContestForm()
        db_sess = db_session.create_session()
        tasks_data = db_sess.query(Task).filter(
            Task.author_id == current_user.id).all()
        contest_data = db_sess.query(Contest).filter(
            Contest.id == contest_id).first()
        if form.validate_on_submit():
            task_id = request.form.get('id')
            task = db_sess.query(Task).filter(Task.id == task_id).first()
            contest_data.tasks.append(task)
            db_sess.commit()
        if not tasks_data:
            return redirect(url_for("tasks_add",
                                    contest_id=contest_id))
        return render_template("tasks.html",
                               title="Список соревнований",
                               tasks=tasks_data,
                               contest=contest_data,
                               form=form)
    return redirect(url_for("index"))


@app.route("/tasks/add", methods=["GET", "POST"])
@login_required
def tasks_add():
    form = AddTasksForContestForm()
    if current_user.job_title == "teacher":
        if form.validate_on_submit():
            db_sess = db_session.create_session()
            input_file = form.task_input.data
            output_file = form.task_output.data
            input_file = str(input_file.read())[
                         2:-1].strip().replace(r"\r\n", "!!!")
            output_file = str(output_file.read())[
                          2:-1].strip().replace(r"\r\n", "!!!")

            task = Task(
                title=form.task_title.data,
                description=form.task_description.data,
                input=input_file,
                output=output_file,
                author_id=current_user.id
            )
            db_sess.add(task)
            db_sess.commit()
            flash('Задание создано!')
            return render_template("add_task_form.html",
                                   title="Добавление задачи",
                                   form=form)
        return render_template("add_task_form.html",
                               title="Добавление задачи",
                               form=form)
    return abort(404)


@app.route("/task_delete/<int:contest_id>/<int:id>")
def task_delete(contest_id, id):
    if current_user.job_title == "teacher":
        db_sess = db_session.create_session()
        task_data = db_sess.query(Task).filter(Task.id == id).filter(Task.author_id == current_user.id).first()
        if task_data:
            contest = db_sess.query(Contest).filter(
                Contest.id == contest_id).first()
            contest.tasks.remove(task_data)
            db_sess.commit()
        else:
            abort(404)
        return redirect(url_for("tasks", contest_id=contest_id))
    return redirect(url_for("index"))


@app.route("/tasks/<int:contest_id>/<int:id>", methods=['GET', 'POST'])
@login_required
def task_edit(contest_id, id):
    if current_user.job_title == "teacher":
        form = AddTasksForContestForm()
        db_sess = db_session.create_session()
        task_data = db_sess.query(Task).filter(Task.id == id).first()
        if request.method == "GET":
            db_sess = db_session.create_session()
            task_data = db_sess.query(Task).filter(Task.id == id,
                                                   Task.author_id == current_user.id).first()
            if task_data:
                form.task_title.data = task_data.title
                form.task_description.data = task_data.description
                form.submit.data = "Обновить"
            else:
                abort(404)
        if form.validate_on_submit():
            db_sess = db_session.create_session()
            task_data = db_sess.query(Task).filter(Contest.id == id,
                                                   Task.author_id == current_user.id).first()
            if task_data:
                task_data.title = form.task_title.data
                task_data.description = form.task_description.data
                task_data.input = str(form.task_input.data.read())[
                                  2:-1].strip().replace(r"\r\n", "!!!")
                task_data.output = str(form.task_output.data.read())[
                                   2:-1].strip().replace(r"\r\n", "!!!")
                db_sess.commit()
                return redirect(url_for("tasks",
                                        contest_id=contest_id))
            else:
                abort(404)

        return render_template("add_task_form.html",
                               title="Редактор задачи",
                               form=form)
    return redirect(url_for("index"))


@app.route("/results/<int:user_id>")
@login_required
def user_results(user_id):
    if current_user.job_title == "student":
        db_sess = db_session.create_session()
        results_data = db_sess.query(ContestResults).filter(ContestResults.student_id == current_user.id).all()
        return render_template("user_results.html",
                               title="Результаты",
                               results=results_data)
    return redirect(url_for("index"))


@app.route("/results")
@login_required
def results():
    db_sess = db_session.create_session()
    contests_data = db_sess.query(Contest).filter(Contest.end_deadline <= dt.datetime.now()).all()
    return render_template("results.html", contests=contests_data)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))


@app.route("/user_avatar")
@login_required
def user_avatar():
    img = current_user.avatar
    if not img:
        with open('static/img/avatar.jpeg', 'rb') as image:
            img = image.read()
    avatar = make_response(img)
    avatar.headers['Content-Type'] = 'image/png'
    return avatar


@app.route("/get_contests_data", methods=["POST"])
@login_required
def get_contest_data():
    if request.method == "POST":
        data = request.get_json()
        db_sess = db_session.create_session()
        if 'count' in data:
            contests_data = db_sess.query(Contest).filter(Contest.id == int(data['contest_id'])
                                                          ,Contest.end_deadline <= dt.datetime.now()).all()
            if not contests_data:
                contest_results = ContestResults(
                    student_id=current_user.id,
                    contest_id=data['contest_id'],
                    complited=data['count'],
                    count_tasks=data['totalCount']
                )
                db_sess.add(contest_results)
                db_sess.commit()

            return jsonify(data)
        return jsonify({"error": "No data."})


@app.route("/download/excel/<int:contest_id>")
@login_required
def download_excel(contest_id):
    db_sess = db_session.create_session()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Результаты")
    ws = wb.active
    cols = ["Фамилия", "Имя", "Отчество", "Результат", "Итог"]
    index_cols = 0
    users_data = db_sess.query(ContestResults).filter(ContestResults.contest_id == contest_id).all()
    for col in ws.iter_cols(min_row=1, max_col=5, max_row=1):
        for cell in col:
            cell.value = cols[index_cols]
            index_cols += 1

    for user_data in enumerate(users_data):
        ws[f'A{user_data[0] + 2}'] = user_data[1].student.surname
        ws[f'B{user_data[0] + 2}'] = user_data[1].student.firstname
        ws[f'C{user_data[0] + 2}'] = user_data[1].student.patronymic
        ws[f'D{user_data[0] + 2}'] = f"{user_data[1].complited}/{user_data[1].count_tasks}"
        if user_data[1].complited == user_data[1].count_tasks:
            ws[f'E{user_data[0] + 2}'] = "Диплом I степени"
        elif int(user_data[1].complited / user_data[1].count_tasks * 100) > 80:
            ws[f'E{user_data[0] + 2}'] = "Диплом II степени"
        elif int(user_data[1].complited / user_data[1].count_tasks * 100) > 60:
            ws[f'E{user_data[0] + 2}'] = "Диплом III степени"
        else:
            ws[f'E{user_data[0] + 2}'] = "Участник"
    wb.save("./static/files/results.xlsx")
    return send_from_directory("./static/files", "results.xlsx", as_attachment=True)


@app.errorhandler(503)
def not_found(error):
    return render_template('503error.html',
                           title="Ошибка 503"), 503


@app.errorhandler(500)
def not_found(error):
    return render_template('500error.html',
                           title="Ошибка 500"), 500


@app.errorhandler(405)
def not_allowed(error):
    return render_template('405error.html',
                           title="Ошибка 405"), 405


@app.errorhandler(404)
def not_found(error):
    return render_template('404error.html',
                           title="Ошибка 404"), 404


@app.errorhandler(403)
def not_found(error):
    return render_template('403error.html',
                           title="Ошибка 403"), 403


@app.errorhandler(401)
def unauthorized(error):
    return render_template('401error.html',
                           title="Ошибка 401"), 401


if __name__ == '__main__':
    app.run(debug=True, port=5050)
